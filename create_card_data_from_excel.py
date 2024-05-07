import json
import openpyxl

def read_excel_to_list(filename):
    # Load the workbook
    wb = openpyxl.load_workbook(filename)
    # Select the active worksheet
    ws = wb.active

    # Initialize an empty list to store dictionaries
    data = []

    # Iterate through rows in the worksheet
    for index, row in enumerate(ws.iter_rows(values_only=True)):

        if index == 0:
            continue

        # Initialize an empty dictionary for each row
        row_dict = {}
        row_dict["index"] = index-1
        # Iterate through cells in the row
        for idx, cell_value in enumerate(row):
            # Get the column header as key for the dictionary
            key = ws.cell(row=1, column=idx+1).value
            # Assign cell value to corresponding key in dictionary
            row_dict[key] = cell_value
        # Append the dictionary to the list
        data.append(row_dict)

    return data

# Example usage
filename = 'cards.xlsx'  # Replace with your file path
data = read_excel_to_list(filename)

with open("extracted_card_data.json",'w') as json_file:
    json.dump(data,json_file,indent=4)
